import streamlit as st
import asyncio
import os
import json
import tempfile
import pandas as pd
from typing import Dict, Any, List, Optional
import time
from datetime import datetime
import traceback
import sys

# 导入PPT更新相关模块
from ppt_parser import PPTParser
from ppt_update_agent_full import PPTUpdateAgent

# 设置页面配置
st.set_page_config(
    page_title="PPT智能更新助手",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义CSS样式
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .section-header {
        font-size: 1.5rem;
        color: #2e8b57;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }
    .update-card {
        background-color: #f0f8ff;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #1f77b4;
        margin: 1rem 0;
    }
    .original-content {
        background-color: #fff5ee;
        padding: 1rem;
        border-radius: 5px;
        border-left: 3px solid #ff6347;
    }
    .updated-content {
        background-color: #f0fff0;
        padding: 1rem;
        border-radius: 5px;
        border-left: 3px solid #32cd32;
    }
    .slide-content {
        background-color: #fafafa;
        padding: 1rem;
        border-radius: 5px;
        border: 1px solid #ddd;
        margin: 0.5rem 0;
    }
    .log-container {
        background-color: #1e1e1e;
        color: #ffffff;
        padding: 1rem;
        border-radius: 5px;
        font-family: 'Courier New', monospace;
        font-size: 12px;
        max-height: 400px;
        overflow-y: auto;
        border: 1px solid #333;
        margin: 1rem 0;
    }
    .log-entry {
        margin: 2px 0;
        padding: 2px 0;
    }
    .log-timestamp {
        color: #888;
    }
    .log-action {
        color: #4CAF50;
        font-weight: bold;
    }
    .log-data {
        color: #FFC107;
    }
</style>
""", unsafe_allow_html=True)

def init_session_state():
    """初始化session state"""
    if 'ppt_parser' not in st.session_state:
        st.session_state.ppt_parser = None
    if 'update_agent' not in st.session_state:
        st.session_state.update_agent = None
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'uploaded_file_name' not in st.session_state:
        st.session_state.uploaded_file_name = None
    if 'slide_contents' not in st.session_state:
        st.session_state.slide_contents = []
    if 'uploaded_file_data' not in st.session_state:
        st.session_state.uploaded_file_data = None


def display_header():
    """显示页面标题"""
    st.markdown('<h1 class="main-header">📊 PPT智能更新助手</h1>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("""
    ### 🚀 功能介绍
    - **智能分析**：自动识别PPT中需要更新的内容
    - **实时搜索**：获取最新信息和数据
    - **内容建议**：生成专业的更新建议
    - **可视化展示**：直观展示原内容与更新内容的对比
    """)

def upload_ppt_file():
    """PPT文件上传功能"""
    st.markdown('<h2 class="section-header">📁 上传PPT文件</h2>', unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader(
        "选择PPT文件",
        type=['pptx'],
        help="支持.pptx格式的PowerPoint文件"
    )
    
    if uploaded_file is not None:
        if st.session_state.uploaded_file_name != uploaded_file.name:
            # 新文件上传，重置状态
            st.session_state.uploaded_file_name = uploaded_file.name
            st.session_state.uploaded_file_data = uploaded_file.getvalue()  # 保存文件数据
            st.session_state.ppt_parser = None
            st.session_state.analysis_results = None
            st.session_state.slide_contents = []

            # 保存上传的文件到临时目录
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pptx') as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_file_path = tmp_file.name

            # 解析PPT文件
            with st.spinner('正在解析PPT文件...'):
                try:
                    parser = PPTParser(tmp_file_path)
                    slide_contents = parser.parse()
                    st.session_state.ppt_parser = parser
                    st.session_state.slide_contents = slide_contents
                    st.success(f'✅ 成功解析PPT文件：{uploaded_file.name}')
                    st.info(f'📄 共发现 {len(slide_contents)} 张幻灯片')
                except Exception as e:
                    st.error(f'❌ 解析PPT文件时出错：{str(e)}')
                    return None
                finally:
                    # 清理临时文件
                    if os.path.exists(tmp_file_path):
                        os.unlink(tmp_file_path)

        return uploaded_file
    
    return None

def display_ppt_content():
    """显示PPT内容"""
    if not st.session_state.slide_contents:
        return
    
    st.markdown('<h2 class="section-header">📖 PPT内容预览</h2>', unsafe_allow_html=True)
    
    # 创建标签页显示不同幻灯片
    slide_tabs = st.tabs([f"幻灯片 {i+1}" for i in range(len(st.session_state.slide_contents))])
    
    for i, tab in enumerate(slide_tabs):
        with tab:
            slide = st.session_state.slide_contents[i]
            
            # 显示文本内容
            if slide['texts']:
                st.markdown("**📝 文本内容：**")
                for j, text_item in enumerate(slide['texts']):
                    text_type = "🏷️ 标题" if text_item['type'] == 'title' else "📄 正文"
                    st.markdown(f'<div class="slide-content"><strong>{text_type}：</strong><br>{text_item["text"]}</div>', 
                              unsafe_allow_html=True)
            
            # 显示表格内容
            if slide['tables']:
                st.markdown("**📊 表格内容：**")
                for j, table in enumerate(slide['tables']):
                    st.markdown(f"表格 {j+1}：")
                    if not table['dataframe'].empty:
                        st.dataframe(table['dataframe'])
                    else:
                        st.write("表格数据为空")
            
            # 显示图片信息
            if slide['images']:
                st.markdown("**🖼️ 图片信息：**")
                for j, image in enumerate(slide['images']):
                    st.markdown(f"- 图片 {j+1}: {image['description']}")
            
            # 显示备注
            if slide['notes']:
                st.markdown("**📝 备注：**")
                st.markdown(f'<div class="slide-content">{slide["notes"]}</div>', unsafe_allow_html=True)

async def initialize_update_agent():
    """初始化更新Agent"""
    if st.session_state.update_agent is None:
        config_file = os.path.join(os.path.dirname(__file__), "browser_mcp.json")
        if not os.path.exists(config_file):
            st.error(f"❌ 配置文件不存在：{config_file}")
            return False
        
        try:
            agent = PPTUpdateAgent(config_file)
            await agent.initialize()
            st.session_state.update_agent = agent
            return True
        except Exception as e:
            st.error(f"❌ 初始化更新Agent失败：{str(e)}")
            return False
    
    return True



def run_async_function(coro):
    """运行异步函数的辅助函数"""
    try:
        # 尝试获取当前事件循环
        loop = asyncio.get_event_loop()
        if loop.is_running():
            # 如果循环正在运行，创建新的事件循环
            import nest_asyncio
            nest_asyncio.apply()
            return asyncio.run(coro)
        else:
            return loop.run_until_complete(coro)
    except RuntimeError:
        # 如果没有事件循环，创建新的
        return asyncio.run(coro)
    except Exception as e:
        st.error(f"❌ 异步函数执行失败：{str(e)}")
        st.error(f"详细错误信息：{traceback.format_exc()}")
        return None

def analysis_control_panel():
    """分析控制面板"""
    if not st.session_state.slide_contents:
        return

    st.markdown('<h2 class="section-header">🔍 智能分析控制</h2>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns([2, 1, 1])

    with col1:
        # 选择要分析的幻灯片
        slide_options = [f"幻灯片 {i+1}" for i in range(len(st.session_state.slide_contents))]
        slide_options.insert(0, "全部幻灯片")

        selected_slides = st.multiselect(
            "选择要分析的幻灯片",
            options=slide_options,
            default=["全部幻灯片"],
            help="选择需要进行智能分析的幻灯片"
        )

    with col2:
        # 分析按钮
        if st.button("🚀 开始分析", type="primary", use_container_width=True):
            if selected_slides:
                # 确定要分析的幻灯片编号
                if "全部幻灯片" in selected_slides:
                    slides_to_analyze = None  # 分析所有幻灯片
                else:
                    slides_to_analyze = [int(slide.split()[1]) for slide in selected_slides]

                # 运行分析
                run_analysis(slides_to_analyze)
            else:
                st.warning("请选择要分析的幻灯片")

    with col3:
        # 清除结果按钮
        if st.button("🗑️ 清除结果", use_container_width=True):
            st.session_state.analysis_results = None
            st.rerun()

def run_analysis(slides_to_analyze):
    """运行PPT分析 - 按照ppt_update_agent_full.py的流程"""
    if not st.session_state.uploaded_file_name:
        st.error("❌ 请先上传PPT文件")
        return

    # 检查API密钥
    if not os.environ.get("DEEPSEEK_API_KEY"):
        st.error("❌ 请在侧边栏设置DeepSeek API密钥")
        return



    # 初始化更新Agent
    with st.spinner('正在初始化分析引擎...'):
        success = run_async_function(initialize_update_agent())
        if not success:
            st.error("❌ 初始化分析引擎失败，请检查配置")
            return

    # 使用保存的文件数据
    if not st.session_state.uploaded_file_data:
        st.error("❌ 无法获取上传的文件数据")
        return

    # 保存上传文件到临时位置
    tmp_file_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pptx') as tmp_file:
            tmp_file.write(st.session_state.uploaded_file_data)
            tmp_file_path = tmp_file.name

        with st.spinner('🔍 正在分析PPT内容，搜索最新信息...'):
            # 显示进度条
            progress_bar = st.progress(0)
            status_text = st.empty()

            status_text.text('正在初始化PPT更新Agent...')
            progress_bar.progress(10)

            # 使用PPTUpdateAgent的process_ppt方法
            status_text.text('正在分析PPT内容...')
            progress_bar.progress(30)

            # 调用实际的分析流程
            try:
                analysis_results = run_async_function(
                    st.session_state.update_agent.process_ppt(tmp_file_path, slides_to_analyze)
                )

                if analysis_results is None:
                    st.error("❌ 分析过程失败，请重试")
                    return

                progress_bar.progress(80)
                status_text.text('正在生成更新建议...')

                # 检查是否有错误
                if "错误" in analysis_results:
                    st.error(f"❌ 分析失败：{analysis_results['错误']}")
                    return

            except Exception as api_error:
                error_msg = str(api_error)
                if "402" in error_msg or "Insufficient Balance" in error_msg:
                    st.error("❌ API余额不足，请充值您的DeepSeek账户后重试")
                    st.info("💡 提示：请登录DeepSeek官网检查账户余额并进行充值")
                elif "401" in error_msg or "Unauthorized" in error_msg:
                    st.error("❌ API密钥无效，请检查您的DeepSeek API密钥")
                elif "429" in error_msg or "rate_limit" in error_msg.lower():
                    st.error("❌ API调用频率超限，请稍后重试")
                else:
                    st.error(f"❌ API调用失败：{error_msg}")
                return



            # 添加分析时间
            analysis_results["分析时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            progress_bar.progress(100)
            status_text.text('分析完成！')

            # 保存结果
            st.session_state.analysis_results = analysis_results

            # 清除进度显示
            progress_bar.empty()
            status_text.empty()

            # 显示分析摘要
            total_updates = sum(len(slide.get("更新内容", [])) for slide in analysis_results.get("结果", []))
            if total_updates > 0:
                st.success(f'✅ 分析完成！发现 {total_updates} 个需要更新的内容项。')
            else:
                st.info('ℹ️ 分析完成！未发现需要更新的内容。')

    except Exception as e:
        st.error(f'❌ 分析过程中出错：{str(e)}')
        with st.expander("查看详细错误信息"):
            st.code(traceback.format_exc())
    finally:
        # 清理临时文件
        if tmp_file_path and os.path.exists(tmp_file_path):
            try:
                os.unlink(tmp_file_path)
            except Exception as e:
                st.warning(f"⚠️ 清理临时文件失败：{str(e)}")


def display_analysis_results():
    """显示分析结果"""
    if not st.session_state.analysis_results:
        return

    st.markdown('<h2 class="section-header">📊 分析结果</h2>', unsafe_allow_html=True)

    results = st.session_state.analysis_results

    # 显示分析摘要
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("分析幻灯片数", len(results.get("处理的幻灯片", [])))
    with col2:
        total_updates = sum(len(slide.get("更新内容", [])) for slide in results.get("结果", []))
        st.metric("发现更新项", total_updates)
    with col3:
        st.metric("分析时间", results.get("分析时间", "未知"))

    # 显示详细结果
    for slide_result in results.get("结果", []):
        slide_num = slide_result.get("幻灯片", 0)

        with st.expander(f"📄 幻灯片 {slide_num} 分析结果", expanded=True):
            # 显示需要更新的部分
            updates = slide_result.get("更新内容", [])

            if updates:
                for i, update in enumerate(updates):
                    st.markdown(f'<div class="update-card">', unsafe_allow_html=True)
                    st.markdown(f"**更新项 {i+1}**")

                    # 原内容
                    st.markdown("**🔴 原内容：**")
                    st.markdown(f'<div class="original-content">{update.get("原内容", "")}</div>',
                              unsafe_allow_html=True)

                    # 更新内容
                    st.markdown("**🟢 建议更新为：**")
                    st.markdown(f'<div class="updated-content">{update.get("更新内容", "")}</div>',
                              unsafe_allow_html=True)

                    # 搜索关键词和结果
                    st.markdown("**🔍 搜索详情：**")
                    st.markdown(f"**搜索关键词：** {update.get('搜索关键词', '')}")
                    st.markdown("**搜索结果：**")
                    st.text_area("", update.get("搜索结果", ""), height=150, key=f"search_{slide_num}_{i}")

                    st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.info("该幻灯片未发现需要更新的内容")

def export_results():
    """导出分析结果"""
    if not st.session_state.analysis_results:
        return

    st.markdown('<h2 class="section-header">📥 导出结果</h2>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)

    with col1:
        # 导出完整JSON格式
        if st.button("📄 导出完整JSON报告", use_container_width=True):
            # 创建完整报告
            complete_report = st.session_state.analysis_results.copy()

            json_data = json.dumps(complete_report, ensure_ascii=False, indent=2)
            st.download_button(
                label="下载完整JSON文件",
                data=json_data,
                file_name=f"ppt_complete_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                key="download_complete_json"
            )

    with col2:
        # 导出更新内容JSON格式
        if st.button("📋 导出更新JSON", use_container_width=True):
            # 创建只包含更新内容的JSON
            update_report = {
                "文件": st.session_state.analysis_results.get("文件", ""),
                "分析时间": st.session_state.analysis_results.get("分析时间", ""),
                "处理的幻灯片": st.session_state.analysis_results.get("处理的幻灯片", []),
                "更新内容": []
            }

            for slide_result in st.session_state.analysis_results.get("结果", []):
                slide_updates = {
                    "幻灯片": slide_result.get("幻灯片", 0),
                    "更新项": slide_result.get("更新内容", [])
                }
                update_report["更新内容"].append(slide_updates)

            json_data = json.dumps(update_report, ensure_ascii=False, indent=2)
            st.download_button(
                label="下载更新JSON文件",
                data=json_data,
                file_name=f"ppt_updates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                key="download_updates_json"
            )

    with col3:
        # 导出Excel格式
        if st.button("📊 导出Excel报告", use_container_width=True):
            # 创建Excel数据
            excel_data = create_excel_report(st.session_state.analysis_results)
            st.download_button(
                label="下载Excel文件",
                data=excel_data,
                file_name=f"ppt_analysis_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel"
            )

def create_excel_report(results):
    """创建Excel报告"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "PPT分析报告"

    # 设置标题
    ws['A1'] = "PPT智能更新分析报告"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')

    # 设置表头
    headers = ['幻灯片', '原内容', '更新内容', '搜索关键词', '更新原因', '分析时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # 填充数据
    row = 4
    for slide_result in results.get("结果", []):
        slide_num = slide_result.get("幻灯片", 0)
        updates = slide_result.get("更新内容", [])

        for update in updates:
            ws.cell(row=row, column=1, value=f"幻灯片 {slide_num}")
            ws.cell(row=row, column=2, value=update.get("原内容", ""))
            ws.cell(row=row, column=3, value=update.get("更新内容", ""))
            ws.cell(row=row, column=4, value=update.get("搜索关键词", ""))
            ws.cell(row=row, column=5, value="内容可能过时")
            ws.cell(row=row, column=6, value=results.get("分析时间", ""))
            row += 1

    # 调整列宽
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()

def sidebar_config():
    """侧边栏配置"""
    with st.sidebar:
        st.markdown("## ⚙️ 配置选项")

        # API配置
        st.markdown("### 🔑 API配置")
        api_key = st.text_input(
            "DeepSeek API Key",
            type="password",
            help="输入您的DeepSeek API密钥"
        )

        if api_key:
            os.environ["DEEPSEEK_API_KEY"] = api_key
            st.success("✅ API密钥已设置")
        else:
            st.warning("⚠️ 请输入DeepSeek API密钥")

        # API使用提示
        with st.expander("💡 API使用提示"):
            st.markdown("""
            **获取API密钥：**
            1. 访问 [DeepSeek官网](https://platform.deepseek.com/)
            2. 注册并登录账户
            3. 在控制台中创建API密钥

            **常见问题：**
            - **余额不足**：请在官网充值账户
            - **密钥无效**：请检查密钥是否正确复制
            - **调用超限**：请稍后重试或升级套餐
            """)

        # 分析参数
        st.markdown("### 🎛️ 分析参数")
        max_steps = st.slider(
            "最大分析步数",
            min_value=1,
            max_value=10,
            value=5,
            help="控制分析的深度和详细程度"
        )

        # 模型选择
        model_options = ["deepseek-chat"]
        selected_model = st.selectbox(
            "选择语言模型",
            options=model_options,
            index=0,
            help="选择用于分析的语言模型"
        )


        # 保存配置
        config = {
            "max_steps": max_steps,
            "model": selected_model,
        }

        return config

def display_statistics():
    """显示统计信息"""
    if not st.session_state.analysis_results:
        return

    st.markdown('<h2 class="section-header">📈 分析统计</h2>', unsafe_allow_html=True)

    results = st.session_state.analysis_results

    # 计算统计数据
    total_slides = len(results.get("处理的幻灯片", []))
    total_updates = sum(len(slide.get("更新内容", [])) for slide in results.get("结果", []))

    # 显示总体统计
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("处理幻灯片数", total_slides)
    with col2:
        st.metric("总更新项数", total_updates)
    with col3:
        avg_updates = round(total_updates / total_slides, 1) if total_slides > 0 else 0
        st.metric("平均更新项/幻灯片", avg_updates)

    # 按幻灯片统计
    slide_stats = []
    for slide_result in results.get("结果", []):
        slide_num = slide_result.get("幻灯片", 0)
        update_count = len(slide_result.get("更新内容", []))
        slide_stats.append({
            "幻灯片": f"幻灯片 {slide_num}",
            "更新项数量": update_count
        })

   

def display_help():
    """显示帮助信息"""
    with st.expander("❓ 使用帮助", expanded=False):
        st.markdown("""
        ### 📖 使用指南

        #### 1. 上传PPT文件
        - 支持 `.pptx` 格式的PowerPoint文件
        - 文件大小建议不超过50MB
        - 确保PPT包含文本内容（图片内容暂不支持分析）

        #### 2. 查看PPT内容
        - 系统会自动解析PPT中的文本、表格和图片信息
        - 可以通过标签页查看不同幻灯片的内容

        #### 3. 配置分析参数
        - 在侧边栏设置API密钥和分析参数
        - 选择要分析的幻灯片范围
        - 调整更新敏感度

        #### 4. 开始智能分析
        - 点击"开始分析"按钮
        - 系统会自动识别需要更新的内容
        - 搜索最新信息并生成更新建议

        #### 5. 查看分析结果
        - 查看原内容与建议更新内容的对比
        - 了解更新原因和搜索依据
        - 查看详细的搜索结果

        #### 6. 导出结果
        - 支持导出JSON和Excel格式的分析报告
        - 包含完整的分析过程和结果

        ### ⚠️ 注意事项
        - 需要有效的DeepSeek API密钥
        - 分析过程需要网络连接
        - 分析时间取决于PPT内容复杂度
        - 建议在分析前检查网络连接状态

        ### 🔧 故障排除
        - 如果上传失败，请检查文件格式和大小
        - 如果分析失败，请检查API密钥和网络连接
        - 如果结果不准确，可以调整分析参数重新分析
        """)

def main():
    """主函数"""
    init_session_state()

    # 侧边栏配置
    sidebar_config()

    # 主界面
    display_header()

    # 帮助信息
    display_help()

    # 上传PPT文件
    uploaded_file = upload_ppt_file()

    if uploaded_file and st.session_state.slide_contents:
        # 显示PPT内容
        display_ppt_content()

        # 分析控制面板
        analysis_control_panel()

        # 显示分析结果
        display_analysis_results()



        # 显示统计信息
        display_statistics()

        # 导出结果
        if st.session_state.analysis_results:
            export_results()

    # 页脚
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; padding: 20px;'>
        <p>📊 PPT智能更新助手 | 基于AI技术的智能内容分析工具</p>
        <p>💡 如有问题或建议，请联系技术支持</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
